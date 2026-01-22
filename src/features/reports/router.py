from flask import Blueprint, request, jsonify, send_file
# from flask_jwt_extended import jwt_required
from flask_jwt_extended import jwt_required, get_jwt_identity
from core.database.connection import db
from .repository import ReporteHornosRepository
from .service import ReporteHornosService
import io
from datetime import datetime

reportes_bp = Blueprint('reportes', __name__, url_prefix='/api/reportes')

@reportes_bp.route('/hornos', methods=['GET', 'POST'])
# @jwt_required()
def obtener_reporte_hornos():
    """
    GET /api/reportes/hornos?fecha_desde=2025-01-01&fecha_hasta=2026-01-31&numero_ot=OT-12345&page=1&size=10
    POST /api/reportes/hornos
    Body JSON:
    {
        "fecha_desde": "2026-01-01",
        "fecha_hasta": "2026-01-31",
        "numero_ot": "OT-12345",
        "page": 1,
        "size": 10
    }
    """
    try:
        # Defaults paginación
        page = None
        size = None

        # Determinar si es GET o POST y obtener parámetros
        if request.method == 'GET':
            fecha_desde = request.args.get('fecha_desde')
            fecha_hasta = request.args.get('fecha_hasta')
            numero_ot = request.args.get('numero_ot')

            # ✅ paginado desde query
            page = request.args.get('page', None)
            size = request.args.get('size', None)

        else:  # POST
            if request.is_json:
                data = request.get_json() or {}
                fecha_desde = data.get('fecha_desde')
                fecha_hasta = data.get('fecha_hasta')
                numero_ot = data.get('numero_ot')

                # ✅ paginado desde body
                page = data.get('page')
                size = data.get('size')
            else:
                # Fallback a query params si no hay JSON
                fecha_desde = request.args.get('fecha_desde')
                fecha_hasta = request.args.get('fecha_hasta')
                numero_ot = request.args.get('numero_ot')

                # ✅ paginado desde query (fallback)
                page = request.args.get('page', None)
                size = request.args.get('size', None)

        # Obtener conexión
        with db.get_connection() as conn:
            repository = ReporteHornosRepository(conn)
            service = ReporteHornosService(repository)

            # ✅ Generar reporte (si page/size vienen => paginado; si no => modo antiguo)
            resultado = service.generar_reporte(
                fecha_desde=fecha_desde,
                fecha_hasta=fecha_hasta,
                numero_ot=numero_ot,
                page=page,
                size=size
            )

            return jsonify(resultado), 200

    except ValueError as e:
        return jsonify({
            "success": False,
            "error": f"Error en formato de fecha: {str(e)}"
        }), 400
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@reportes_bp.route('/hornos/excel', methods=['POST'])
def exportar_reporte_hornos_excel():
    """
    POST /api/reportes/hornos/excel
    Genera Excel profesional con diseño limpio y bien distribuido
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        print("✅ Generando Excel...")

        # Obtener parámetros del body
        data = request.get_json() if request.is_json else {}
        fecha_desde = data.get('fecha_desde')
        fecha_hasta = data.get('fecha_hasta')
        numero_ot = data.get('numero_ot')

        # Obtener datos (✅ aquí NO se pagina; exporta todo con los filtros)
        with db.get_connection() as conn:
            repository = ReporteHornosRepository(conn)
            service = ReporteHornosService(repository)

            resultado = service.generar_reporte(
                fecha_desde=fecha_desde,
                fecha_hasta=fecha_hasta,
                numero_ot=numero_ot
            )

        if not resultado.get('success') or not resultado.get('data'):
            return jsonify({
                "success": False,
                "error": "No hay datos para exportar"
            }), 404

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Hornos"

        # ==========================================
        # ESTILOS PROFESIONALES
        # ==========================================

        header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        header_fill = PatternFill(start_color="2C5F8D", end_color="2C5F8D", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        data_alignment = Alignment(horizontal="left", vertical="center")
        data_font = Font(size=10, name="Calibri")
        number_alignment = Alignment(horizontal="right", vertical="center")

        # ==========================================
        # TÍTULO DEL REPORTE
        # ==========================================
        ws.merge_cells('A1:S1')
        title_cell = ws['A1']
        title_cell.value = f"REPORTE DE HORNOS - MODEPSA"
        title_cell.font = Font(bold=True, size=14, color="2C5F8D", name="Calibri")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells('A2:S2')
        subtitle_cell = ws['A2']
        fecha_reporte = datetime.now().strftime('%d/%m/%Y %H:%M')
        subtitle_cell.value = f"Generado el: {fecha_reporte}"
        subtitle_cell.font = Font(size=10, color="666666", name="Calibri")
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[3].height = 5

        # ==========================================
        # HEADERS DE COLUMNAS
        # ==========================================
        headers = [
            ("Fecha Registro", 20),
            ("Número OT", 15),
            ("Tiempo Asignado.", 12),
            ("Peso Total", 12),
            ("Fecha Fin Manual", 18),
            ("Fecha Fin Auto", 18),
            ("Modo Ingreso", 14),
            ("Dureza 1", 10),
            ("Dureza 2", 10),
            ("Dureza 3", 10),
            ("Fecha Modif.", 18),
            ("Usuario", 12),
            ("Temp H1", 10),
            ("Temp H2", 10),
            ("Temp H3", 10),
            ("Temp H4", 10),
            ("Temp H5", 10),
            ("Temp H6", 10),
            ("Temp H7", 10)
        ]

        header_row = 4

        for col_num, (header_text, width) in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header_text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = width

        ws.row_dimensions[header_row].height = 30

        # ==========================================
        # ESCRIBIR DATOS
        # ==========================================
        for row_num, registro in enumerate(resultado['data'], header_row + 1):
            if row_num % 2 == 0:
                row_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            else:
                row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            cell = ws.cell(row=row_num, column=1)
            cell.value = str(registro.get('Fecha_Registro')) if registro.get('Fecha_Registro') else "-"
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            cell.fill = row_fill

            cell = ws.cell(row=row_num, column=2)
            cell.value = registro.get('Numero_OT') or "-"
            cell.font = Font(bold=True, size=10, color="2C5F8D", name="Calibri")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.fill = row_fill

            cell = ws.cell(row=row_num, column=3)
            peso_unit = registro.get('Peso_Unitario')
            cell.value = float(peso_unit) if peso_unit else None
            cell.number_format = '#,##0.00'
            cell.font = data_font
            cell.alignment = number_alignment
            cell.border = thin_border
            cell.fill = row_fill

            cell = ws.cell(row=row_num, column=4)
            peso_total = registro.get('Peso_Total')
            cell.value = float(peso_total) if peso_total else None
            cell.number_format = '#,##0.00'
            cell.font = Font(bold=True, size=10, color="28A745", name="Calibri")
            cell.alignment = number_alignment
            cell.border = thin_border
            cell.fill = row_fill

            cell = ws.cell(row=row_num, column=5)
            cell.value = str(registro.get('Fecha_Fin_Manual')) if registro.get('Fecha_Fin_Manual') else "-"
            cell.font = Font(size=9, name="Calibri")
            cell.alignment = data_alignment
            cell.border = thin_border
            cell.fill = row_fill

            cell = ws.cell(row=row_num, column=6)
            cell.value = str(registro.get('Fecha_Fin_Auto')) if registro.get('Fecha_Fin_Auto') else "-"
            cell.font = Font(size=9, name="Calibri")
            cell.alignment = data_alignment
            cell.border = thin_border
            cell.fill = row_fill

            cell = ws.cell(row=row_num, column=7)
            modo = registro.get('Modo_Ingreso_Carga')
            cell.value = modo or "-"
            if modo == "Automatico":
                cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                cell.font = Font(bold=True, size=10, color="155724", name="Calibri")
            else:
                cell.fill = row_fill
                cell.font = data_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

            for col_offset, dureza_key in enumerate(['Dureza_1', 'Dureza_2', 'Dureza_3']):
                cell = ws.cell(row=row_num, column=8 + col_offset)
                dureza = registro.get(dureza_key)
                cell.value = float(dureza) if dureza else None
                cell.number_format = '#,##0.0'
                cell.font = data_font
                cell.alignment = number_alignment
                cell.border = thin_border
                cell.fill = row_fill

            cell = ws.cell(row=row_num, column=11)
            cell.value = str(registro.get('Fecha_Modificacion')) if registro.get('Fecha_Modificacion') else "-"
            cell.font = Font(size=9, name="Calibri")
            cell.alignment = data_alignment
            cell.border = thin_border
            cell.fill = row_fill

            cell = ws.cell(row=row_num, column=12)
            cell.value = registro.get('Usuario') or "-"
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.fill = row_fill

            temp_keys = [
                'TEMPERATURA HORNO1', 'TEMPERATURA HORNO2', 'TEMPERATURA HORNO3',
                'TEMPERATURA HORNO4', 'TEMPERATURA HORNO5', 'TEMPERATURA HORNO6',
                'TEMPERATURA HORNO7'
            ]

            for col_offset, temp_key in enumerate(temp_keys):
                cell = ws.cell(row=row_num, column=13 + col_offset)
                temp = registro.get(temp_key)

                if temp:
                    cell.value = float(temp)
                    cell.number_format = '#,##0.0'
                    cell.font = Font(bold=True, size=10, color="E67E22", name="Calibri")
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                else:
                    cell.value = "-"
                    cell.font = Font(size=9, color="999999", name="Calibri")
                    cell.fill = row_fill

                cell.alignment = number_alignment
                cell.border = thin_border

            ws.row_dimensions[row_num].height = 20

        last_row = len(resultado['data']) + header_row + 2

        ws.merge_cells(f'A{last_row}:S{last_row}')
        footer_cell = ws[f'A{last_row}']
        footer_cell.value = f"Total de registros: {len(resultado['data'])}"
        footer_cell.font = Font(bold=True, size=10, color="666666", name="Calibri")
        footer_cell.alignment = Alignment(horizontal="right", vertical="center")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Reporte_Hornos_{timestamp}.xlsx'

        print(f"✅ Excel generado: {filename}")

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"❌ Error al exportar Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500
